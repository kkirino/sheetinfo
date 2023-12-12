import os
import argparse
import xlrd
import openpyxl
import sys


def config_args():
    parser = argparse.ArgumentParser(
        description="get sheet information in xls/xlsx files without MS Excel application"
    )
    parser.add_argument(
        "-o", "--old", action="store_true", help="input the file in '.xls' format"
    )
    parser.add_argument("excelfile", help="path to excel file")
    return parser.parse_args()


def main():
    args = config_args()
    file_path = os.path.join(os.getcwd(), args.excelfile)

    if args.old:
        try:
            wb = xlrd.open_workbook(file_path)
            print("Name Rows Columns")
            for name in wb.sheet_names():
                sheet = wb.sheet_by_name(name)
                print("{} {} {}".format(name, sheet.nrows, sheet.ncols))
            sys.exit(0)
        except FileNotFoundError as e:
            print("FileNotFoundError: {}".format(e), file=sys.stderr)
            sys.exit(1)
        except xlrd.biffh.XLRDError as e:
            print("xlrd.biffh.XLRDError: {}".format(e), file=sys.stderr)
            sys.exit(1)

    else:
        try:
            wb = openpyxl.load_workbook(file_path)
            print("Name Rows Columns")
            for name in wb.sheetnames:
                sheet = wb[name]
                print("{} {} {}".format(name, sheet.max_row, sheet.max_column))
            sys.exit(0)
        except FileNotFoundError as e:
            print("FileNotFoundError: {}".format(e), file=sys.stderr)
            sys.exit(1)
        except openpyxl.utils.exceptions.InvalidFileException as e:
            print(
                "openpyxl.utils.exceptions.InvalidFileException: {}".format(e),
                file=sys.stderr,
            )
            sys.exit(1)


if __name__ == "__main__":
    main()
